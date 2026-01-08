"""
Views for warehouse API.
"""
from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import LimitOffsetPagination
from rest_framework.response import Response
from rest_framework.views import APIView
from django.views import View
from django.http import HttpResponse, Http404
from django.template.loader import render_to_string
from weasyprint import HTML
from django.db.models import Subquery
from datetime import datetime
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .serializers import *
from core.models import *


class PersonalizedPagination(LimitOffsetPagination):
    default_limit = 10
    max_limit = 100

    def get_paginated_response(self, data):
        return Response({
            'rows': data,
            'total': self.count
        })


class CatalogView(APIView):
    def get(self, request, *args, **kwargs):
        data = []
        selling_channel_id = request.query_params.get("selling_channel_id")
        agency_id = request.query_params.get("agency_id")
        sale_id = request.query_params.get("sale_id")
        purchase_id = request.query_params.get("purchase_id")
        search = request.query_params.get("search", "").strip()

        if selling_channel_id:
            prices = ProductChannelPrice.objects.filter(selling_channel=selling_channel_id)
            products_stock = ProductStock.objects.filter(product__in=Subquery(prices.values('product'))).select_related('product')
            product_price = {
                pp.product_id: pp
                for pp in ProductChannelPrice.objects.filter(selling_channel=selling_channel_id)
            }

            for product_stock in products_stock:
                pp = product_price.get(product_stock.product.id)

                # Skip products without stock records
                if pp is None:
                    continue

                # Apply search filter
                if search and not (search.lower() in product.name.lower() or search.lower() in product.code.lower()):
                    continue

                data.append({
                    "id": product_stock.id,
                    "agency": product_stock.warehouse.agency.name,
                    "warehouse": product_stock.warehouse.name,
                    "name": product_stock.product.name,
                    "code": product_stock.product.code,
                    "price": pp.price,
                    "stock": product_stock.available_stock,
                    "minimum_stock": product_stock.minimum_stock,
                    "maximum_stock": product_stock.maximum_stock,
                    "name": product_stock.product.name,
                    "minimum_sale_price": product_stock.product.minimum_sale_price,
                    "maximum_sale_price": product_stock.product.maximum_sale_price,
                })
        elif agency_id and sale_id:
            sale_items = SaleItem.objects.filter(sale=sale_id, sale__agency=agency_id).exclude(status='completado').prefetch_related('product_stock')

            for sale_item in sale_items:
                product = sale_item.product_stock.product
                
                # Apply search filter
                if search and not (search.lower() in product.name.lower() or search.lower() in product.code.lower()):
                    continue

                data.append({
                    "sale_item_id": sale_item.id,
                    "id": sale_item.product_stock.id,
                    "agency": sale_item.product_stock.warehouse.agency.name,
                    "warehouse": sale_item.product_stock.warehouse.name,
                    "name": product.name,
                    "code": product.code,
                    "price": 0,
                    "stock": sale_item.product_stock.available_stock,
                    "minimum_stock": sale_item.product_stock.minimum_stock,
                    "maximum_stock": sale_item.product_stock.maximum_stock,
                    "minimum_sale_price": product.minimum_sale_price,
                    "maximum_sale_price": product.maximum_sale_price,
                    "status": sale_item.status,
                })
        elif agency_id and purchase_id:
            purchase_items = PurchaseItem.objects.filter(purchase=purchase_id, purchase__agency=agency_id).exclude(status='completado').prefetch_related('product_stock')

            for purchase_item in purchase_items:
                product = purchase_item.product_stock.product
                
                # Apply search filter
                if search and not (search.lower() in product.name.lower() or search.lower() in product.code.lower()):
                    continue

                data.append({
                    "purchase_item_id": purchase_item.id,
                    "id": purchase_item.product_stock.id,
                    "agency": purchase_item.product_stock.warehouse.agency.name,
                    "warehouse": purchase_item.product_stock.warehouse.name,
                    "name": product.name,
                    "code": product.code,
                    "price": 0,
                    "stock": purchase_item.product_stock.available_stock,
                    "minimum_stock": purchase_item.product_stock.minimum_stock,
                    "maximum_stock": purchase_item.product_stock.maximum_stock,
                    "minimum_sale_price": product.minimum_sale_price,
                    "maximum_sale_price": product.maximum_sale_price,
                    "status": purchase_item.status,
                })
        elif agency_id:
            products = ProductStock.objects.filter(warehouse__agency=agency_id).select_related('product')

            for product in products:
                # Apply search filter
                if search and not (search.lower() in product.product.name.lower() or search.lower() in product.product.code.lower()):
                    continue

                data.append({
                    "id": product.id,
                    "agency": product.warehouse.agency.name,
                    "warehouse": product.warehouse.name,
                    "name": product.product.name,
                    "code": product.product.code,
                    "price": 0,
                    "stock": product.stock,
                    "minimum_stock": product.minimum_stock,
                    "maximum_stock": product.maximum_stock,
                    "minimum_sale_price": product.product.minimum_sale_price,
                    "maximum_sale_price": product.product.maximum_sale_price,
                })
        else:
            products = ProductStock.objects.all().select_related('product')

            for product in products:
                # Apply search filter
                if search and not (search.lower() in product.product.name.lower() or search.lower() in product.product.code.lower()):
                    continue

                data.append({
                    "id": product.id,
                    "agency": product.warehouse.agency.name,
                    "warehouse": product.warehouse.name,
                    "name": product.product.name,
                    "code": product.product.code,
                    "price": 0,
                    "stock": product.stock,
                    "minimum_stock": product.minimum_stock,
                    "maximum_stock": product.maximum_stock,
                    "minimum_sale_price": product.product.minimum_sale_price,
                    "maximum_sale_price": product.product.maximum_sale_price,
                })

        return Response(CatalogProductSerializer(data, many=True).data)


class AgencyViewSet(viewsets.ModelViewSet):
    """View for managing agency APIs."""
    serializer_class = AgencySerializer
    queryset = Agency.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'name', 'city']
    pagination_class = PersonalizedPagination

    def get_queryset(self):
        """Retrieve agencies ordered by id."""
        return self.queryset.order_by('-id')
    
    @action(detail=False, methods=['get'])
    def choices(self, request):
        """Get city choices."""
        choices = [
            {'value': choice[0], 'label': choice[1]} 
            for choice in Agency.CITY_CHOICES
        ]
        return Response(choices)
    
    @action(detail=False, methods=["get"], url_path="all")
    def all_agencies(self, request):
        queryset = self.filter_queryset(self.get_queryset()).values("id", "name")
        return Response(list(queryset))


class WarehouseViewSet(viewsets.ModelViewSet):
    """View for managing warehouse APIs."""
    serializer_class = WarehouseSerializer
    queryset = Warehouse.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'name', 'agency__name']
    pagination_class = PersonalizedPagination

    def get_queryset(self):
        """Retrieve warehouses ordered by id."""
        return self.queryset.order_by('-id')


class CategoryViewSet(viewsets.ModelViewSet):
    """View for managing category APIs."""
    serializer_class = CategorySerializer
    queryset = Category.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'name']
    pagination_class = PersonalizedPagination

    def get_queryset(self):
        """Retrieve categories ordered by id."""
        return self.queryset.order_by('-id')
    
    @action(detail=False, methods=["get"], url_path="all")
    def all_categories(self, request):
        queryset = self.filter_queryset(self.get_queryset()).values("id", "name")
        return Response(list(queryset))


class BatchViewSet(viewsets.ModelViewSet):
    """View for managing batch APIs."""
    serializer_class = BatchSerializer
    queryset = Batch.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'name']
    pagination_class = PersonalizedPagination

    def get_queryset(self):
        """Retrieve batches ordered by id."""
        return self.queryset.order_by('-id')
    
    @action(detail=False, methods=["get"], url_path="all")
    def all_categories(self, request):
        queryset = self.filter_queryset(self.get_queryset()).values("id", "name")
        return Response(list(queryset))


class ProductViewSet(viewsets.ModelViewSet):
    """View for managing product APIs."""
    serializer_class = ProductSerializer
    queryset = Product.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'name', 'code']
    pagination_class = PersonalizedPagination

    def get_queryset(self):
        """Retrieve products ordered by id."""
        return self.queryset.order_by('-id')
    
    def update(self, request, *args, **kwargs):
        """Custom update method to handle image deletion."""
        instance = self.get_object()
        
        # Check if image should be deleted (not present in request data or explicitly set to null/empty)
        should_delete_image = (
            'image' not in request.data or 
            request.data.get('image') is None or 
            request.data.get('image') == ''
        )
        
        if should_delete_image and instance.image:
            # Delete the old image file from storage
            instance.delete_image()
        
        return super().update(request, *args, **kwargs)


class ProductStockViewSet(viewsets.ModelViewSet):
    """View for managing product stock APIs."""
    serializer_class = ProductStockSerializer
    queryset = ProductStock.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'product__name', 'product__code', 'warehouse__name']
    pagination_class = PersonalizedPagination

    def get_queryset(self):
        """Retrieve product stocks ordered by id."""
        return self.queryset.order_by('-id')
    
    @action(detail=True, methods=['post'], url_path='increment-damaged-stock')
    def increment_damaged_stock(self, request, pk=None):
        """Increment damaged stock by a given quantity."""
        product_stock = self.get_object()
        serializer = IncrementDamagedStockSerializer(
            data=request.data,
            context={'product_stock': product_stock}
        )
        
        if serializer.is_valid():
            quantity = serializer.validated_data['quantity']
            product_stock.damaged_stock += quantity
            product_stock.available_stock -= quantity
            product_stock.save(update_fields=['damaged_stock', 'available_stock'])
            
            return Response({
                'message': f'Stock dañado incrementado en {quantity} unidades.',
                'damaged_stock': product_stock.damaged_stock
            })
        
        return Response(serializer.errors, status=400)


class ClientViewSet(viewsets.ModelViewSet):
    """View for managing client APIs."""
    serializer_class = ClientSerializer
    queryset = Client.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['id', 'name', 'phone', 'email','nit']
    filterset_fields = ['client_type']
    pagination_class = PersonalizedPagination

    def get_queryset(self):
        """Retrieve clients ordered by id."""
        return self.queryset.order_by('-id')

    @action(detail=False, methods=['get'])
    def choices(self, request):
        """Get client type choices."""
        choices = [
            {'value': choice[0], 'label': choice[1]} 
            for choice in Client.CLIENT_TYPE_CHOICES
        ]
        return Response(choices)
    
    @action(detail=False, methods=["get"], url_path="all")
    def all_clients(self, request):
        queryset = self.filter_queryset(self.get_queryset()).values("id", "name")
        return Response(list(queryset))


class SupplierViewSet(viewsets.ModelViewSet):
    """View for managing supplier APIs."""
    serializer_class = SupplierSerializer
    queryset = Supplier.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'name', 'phone', 'nit']
    pagination_class = PersonalizedPagination

    def get_queryset(self):
        """Retrieve suppliers ordered by id."""
        return self.queryset.order_by('-id')
    
    @action(detail=False, methods=["get"], url_path="all")
    def all_suppliers(self, request):
        queryset = self.filter_queryset(self.get_queryset()).values("id", "name")
        return Response(list(queryset))


class EntryViewSet(viewsets.ModelViewSet):
    """View for managing entry APIs."""
    serializer_class = EntrySerializer
    queryset = Entry.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'invoice_number',
                     'warehouse_keeper__first_name', 'warehouse_keeper__last_name', 'supplier__name']
    pagination_class = PersonalizedPagination
    
    def perform_create(self, serializer):
        serializer.save(warehouse_keeper=self.request.user)

    def get_queryset(self):
        """Retrieve entries ordered by id."""
        return self.queryset.order_by('-id')


class OutputViewSet(viewsets.ModelViewSet):
    """View for managing output APIs."""
    serializer_class = OutputSerializer
    queryset = Output.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter]
    search_fields = ['id',
                     'warehouse_keeper__first_name', 'warehouse_keeper__last_name', 'client__name']
    pagination_class = PersonalizedPagination
    
    def perform_create(self, serializer):
        serializer.save(warehouse_keeper=self.request.user)

    def get_queryset(self):
        """Retrieve outputs ordered by id."""
        return self.queryset.order_by('-id')


class ProductChannelPriceViewSet(viewsets.ModelViewSet):
    """View for managing product channel price APIs."""
    serializer_class = ProductChannelPriceSerializer
    queryset = ProductChannelPrice.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter]
    search_fields = ['product__name', 'selling_channel__name']
    pagination_class = PersonalizedPagination

    def get_queryset(self):
        """Retrieve product channel prices ordered by id."""
        return self.queryset.order_by('-id')


class SellingChannelViewSet(viewsets.ModelViewSet):
    """View for managing selling channel APIs."""
    serializer_class = SellingChannelSerializer
    queryset = SellingChannel.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'name']
    pagination_class = PersonalizedPagination

    def get_queryset(self):
        """Retrieve selling channels ordered by id."""
        return self.queryset.order_by('-id')
    
    @action(detail=False, methods=["get"], url_path="all")
    def all_selling_channels(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class PurchaseViewSet(viewsets.ModelViewSet):
    """View for managing purchase APIs."""
    serializer_class = PurchaseSerializer
    queryset = Purchase.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['id', 'buyer__first_name', 'buyer__last_name', 'supplier__name', 'invoice_number']
    filterset_fields = ['purchase_type', 'status', 'purchase_date']
    pagination_class = PersonalizedPagination

    def perform_create(self, serializer):
        serializer.save(buyer=self.request.user)

    def get_queryset(self):
        """Retrieve purchases ordered by id and filtered by status."""
        queryset = super().get_queryset()
        status = self.request.query_params.get('status')

        if status:
            queryset = queryset.filter(status=status).order_by('-id')

        return self.queryset.order_by('-id')


class SaleViewSet(viewsets.ModelViewSet):
    """View for managin Sale APIs."""
    serializer_class = SaleSerializer
    queryset = Sale.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['id', 'client__name', 'selling_channel__name', 'seller__first_name', 'seller__last_name']
    filterset_fields = ['sale_type', 'status', 'sale_date']
    pagination_class = PersonalizedPagination

    def perform_create(self, serializer):
        serializer.save(seller=self.request.user)

    def get_queryset(self):
        """Retrieve sales ordered by id and filtered by status."""
        queryset = super().get_queryset()
        status = self.request.query_params.get("status")

        if status:
            queryset = queryset.filter(status=status).order_by('-id')
        return self.queryset.order_by('-id')


class PaymentViewSet(viewsets.ModelViewSet):
    """View for managing Payment APIs."""
    serializer_class = PaymentSerializer
    queryset = Payment.objects.all()
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put']
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['id', 'payment_method']
    filterset_fields = ['transaction_type', 'payment_date']
    pagination_class = PersonalizedPagination

    def get_queryset(self):
        """Retrieve payments ordered by id."""
        return self.queryset.order_by('-id')

    
class InvoicePdfView(View):
    """Generate Proforma PDF."""
    def get(self, request, *args, **kwargs):
        sale_id = self.kwargs.get('id')
        try:
            sale = Sale.objects.get(id=sale_id)
        except Sale.DoesNotExist:
            raise Http404("Venta no encontrada.")

        isSale = True if sale.status == 'realizado' else False
        title = 'Proforma' if sale.status == 'proforma' else 'Recibo'

        context = {
            'title': title,
            'sale': sale,
            'isSale': isSale,
        }
        html_string = render_to_string('invoice.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))

        pdf_file = html.write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="comprobante_de_venta_{sale_id}.pdf"'

        return response
    
class OutputInvoicePdfView(View):
    """Generate Output PDF."""
    def get(self, request, *args, **kwargs):
        output_id = self.kwargs.get('id')
        try:
            output = Output.objects.get(id=output_id)
        except Output.DoesNotExist:
            raise Http404("Venta no encontrada.")

        isOutputDone = True if output.sale.status == 'terminado' else False

        context = {
            'title': 'Recibo de Salida',
            'output': output,
            'isOutputDone': isOutputDone,
        }
        html_string = render_to_string('output_invoice.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))

        pdf_file = html.write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="comprobante_de_salida_{output_id}.pdf"'

        return response


class BuyReportPdfView(View):
    """Generate Buy Report PDF."""
    def get(self, request, *args, **kwargs):
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")
        today = datetime.now().date()
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        try:
            purchase_items = PurchaseItem.objects.filter(purchase__purchase_date__range=(start_date, end_date)).select_related("purchase").order_by("purchase__id")
        except PurchaseItem.DoesNotExist:
            raise Http404("Compras no encontradas.")

        context = {
            'title': 'Reporte de Compras',
            'purchase_items': purchase_items,
            'start_date': start_date,
            'end_date': end_date,
            'today': today,
        }
        html_string = render_to_string('buy_report.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))

        pdf_file = html.write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="reporte_de_compras_{start_date}_a_{end_date}.pdf"'

        return response

    
class SellReportPdfView(View):
    """Generate Sell Report PDF."""
    def get(self, request, *args, **kwargs):
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        today = datetime.now().date()
        try:
            sale_items = SaleItem.objects.filter(sale__sale_date__range=(start_date, end_date)).select_related("sale").order_by("sale__id")
        except SaleItem.DoesNotExist:
            raise Http404("Ventas no encontradas.")

        context = {
            'title': 'Reporte de Ventas',
            'sale_items': sale_items,
            'start_date': start_date,
            'end_date': end_date,
            'today': today,
        }
        html_string = render_to_string('sell_report.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))

        pdf_file = html.write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="reporte_de_ventas_{start_date}_a_{end_date}.pdf"'

        return response


class EntryReportPdfView(View):
    """Generate Entry Report PDF."""
    def get(self, request, *args, **kwargs):
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        today = datetime.now().date()
        try:
            entry_items = EntryItem.objects.filter(entry__entry_date__range=(start_date, end_date)).select_related("entry").order_by("entry__id")
        except EntryItem.DoesNotExist:
            raise Http404("Entradas no encontradas.")

        context = {
            'title': 'Reporte de Entradas',
            'entry_items': entry_items,
            'start_date': start_date,
            'end_date': end_date,
            'today': today,
        }
        html_string = render_to_string('entry_report.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))

        pdf_file = html.write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="reporte_de_entradas_{start_date}_a_{end_date}.pdf"'

        return response


class OutputReportPdfView(View):
    """Generate Output Report PDF."""
    def get(self, request, *args, **kwargs):
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        today = datetime.now().date()
        try:
            outputs_items = OutputItem.objects.filter(output__output_date__range=(start_date, end_date)).select_related("output").order_by("output__id")
        except Output.DoesNotExist:
            raise Http404("Salidas no encontradas.")

        context = {
            'title': 'Reporte de Salidas',
            'outputs_items': outputs_items,
            'start_date': start_date,
            'end_date': end_date,
            'today': today,
        }
        html_string = render_to_string('output_report.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))

        pdf_file = html.write_pdf()

        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="reporte_de_salidas_{start_date}_a_{end_date}.pdf"'

        return response


class BuyReportExcelView(APIView):
    """Generate Excel Report."""
    def get(self, request):
        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        try:
            purchase_items = PurchaseItem.objects.filter(purchase__purchase_date__range=(start_date, end_date)).select_related("purchase").order_by("purchase__id")
        except PurchaseItem.DoesNotExist:
            raise Http404("Compras no encontradas.")
        wb = Workbook()
        ws = wb.active
        start_date_formatted = start_date.strftime("%d-%m-%y")
        end_date_formatted = end_date.strftime("%d-%m-%y")
        ws.title = f"Compras {start_date_formatted} - {end_date_formatted}"[:31]
        ws.append(["REPORTE DE COMPRAS DEL " + start_date_formatted + " AL " + end_date_formatted])
        headers = (["AGENCIA",
                   "COMPRADOR",
                   "PROVEEDOR",
                   "TIPO DE COMPRA",
                   "FECHA DE COMPRA",
                   "N° DE FACTURA",
                   "TOTAL",
                   "SALDO PENDIENTE",
                   "PRODUCTO",
                   "CANTIDAD",
                   "PRECIO UNITARIO",
                   "SUB TOTAL",
                   "ESTADO (INGRESO ALMACEN)",
                   ])
        # Unir todas las celdas de la fila 1 (título) y centrarlo
        num_columns = len(headers)
        ws.merge_cells(f'A1:{chr(64 + num_columns)}1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, size=14)
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        column_widths = [
            20,
            25,
            25,
            20,
            18,
            18,
            15,
            18,
            30,
            12,
            18,
            18,
            30,
        ]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        for purchase_item in purchase_items:
            ws.append([
                purchase_item.purchase.agency.name,
                purchase_item.purchase.buyer.first_name + " " + purchase_item.purchase.buyer.last_name,
                purchase_item.purchase.supplier.name,
                purchase_item.purchase.purchase_type,
                purchase_item.purchase.purchase_date,
                purchase_item.purchase.invoice_number,
                purchase_item.purchase.total,
                purchase_item.purchase.balance_due,
                purchase_item.product_stock.product.name,
                purchase_item.quantity,
                purchase_item.unit_price,
                purchase_item.total_price,
                purchase_item.status,
            ])
            
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'inline; filename="reporte_de_compras_{start_date_str}_a_{end_date_str}.xlsx"'

        wb.save(response)
        return response


class SaleReportExcelView(APIView):
    """Generate Excel Report."""
    def get(self, request):
        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        try:
            sale_items = SaleItem.objects.filter(sale__sale_date__range=(start_date, end_date)).select_related("sale").order_by("sale__id")
        except SaleItem.DoesNotExist:
            raise Http404("Ventas no encontradas.")
        wb = Workbook()
        ws = wb.active
        start_date_formatted = start_date.strftime("%d-%m-%y")
        end_date_formatted = end_date.strftime("%d-%m-%y")
        ws.title = f"Ventas {start_date_formatted} - {end_date_formatted}"[:31]
        ws.append(["REPORTE DE VENTAS DEL " + start_date_formatted + " AL " + end_date_formatted])
        headers = (["AGENCIA",
                   "VENDEDOR",
                   "CLIENTE",
                   "CANAL DE VENTA",
                   "TIPO DE VENTA",
                   "FECHA DE VENTA",
                   "N° DE FACTURA",
                   "TOTAL",
                   "SALDO PENDIENTE",   
                   "PRODUCTO",
                   "CANTIDAD",
                   "PRECIO UNITARIO",
                   "SUB TOTAL",
                   "DESCUENTO %",
                   "TOTAL ITEM",
                   "ESTADO (ENTREGA AL CLIENTE)",
                   ])
        # Unir todas las celdas de la fila 1 (título) y centrarlo
        num_columns = len(headers)
        ws.merge_cells(f'A1:{chr(64 + num_columns)}1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, size=14)
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        column_widths = [
            20, # AGENCIA
            25, # VENDEDOR
            25, # CLIENTE
            20, # CANAL DE VENTA
            18, # TIPO DE VENTA
            18, # FECHA DE VENTA    
            15, # N° DE FACTURA
            18, # TOTAL
            30, # SALDO PENDIENTE
            25, # PRODUCTO
            18, # CANTIDAD
            18, # PRECIO UNITARIO
            15, # SUB TOTAL
            20, # DESCUENTO %
            15, # TOTAL ITEM
            30, # ESTADO (ENTREGA AL CLIENTE)
        ]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        previous_sale_id = None
        for sale_item in sale_items:
            current_sale_id = sale_item.sale.id
            # Agregar fila en blanco si cambió el sale_id (excepto en la primera iteración)
            if previous_sale_id is not None and current_sale_id != previous_sale_id:
                ws.append([])
            previous_sale_id = current_sale_id
            
            ws.append([
                sale_item.sale.agency.name,
                sale_item.sale.seller.first_name + " " + sale_item.sale.seller.last_name,
                sale_item.sale.client.name,
                sale_item.sale.selling_channel.name,
                sale_item.sale.sale_type,
                sale_item.sale.sale_date,
                sale_item.sale.invoice_number,
                sale_item.sale.total,
                sale_item.sale.balance_due,
                sale_item.product_stock.product.name,
                sale_item.quantity,
                sale_item.unit_price,
                sale_item.sub_total_price,
                sale_item.discount,
                sale_item.total_price,
                sale_item.status,
            ])
            
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'inline; filename="reporte_de_ventas_{start_date_str}_a_{end_date_str}.xlsx"'

        wb.save(response)
        return response
    

class EntryReportExcelView(APIView):
    """Generate Excel Report."""
    def get(self, request):
        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        try:
            entry_items = EntryItem.objects.filter(entry__entry_date__range=(start_date, end_date)).select_related("entry").order_by("entry__id")
        except EntryItem.DoesNotExist:
            raise Http404("Entradas no encontradas.")
        wb = Workbook()
        ws = wb.active
        start_date_formatted = start_date.strftime("%d-%m-%y")
        end_date_formatted = end_date.strftime("%d-%m-%y")
        ws.title = f"Entradas {start_date_formatted} - {end_date_formatted}"[:31]
        ws.append(["REPORTE DE ENTRADAS DEL " + start_date_formatted + " AL " + end_date_formatted])
        headers = (["AGENCIA",
                   "ALMACENERO",
                   "PROVEEDOR",
                   "FECHA DE ENTRADA",
                   "N° DE RECIBO",
                   "PRODUCTO",
                   "CANTIDAD",
                   ])
        # Unir todas las celdas de la fila 1 (título) y centrarlo
        num_columns = len(headers)
        ws.merge_cells(f'A1:{chr(64 + num_columns)}1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, size=14)
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        column_widths = [
            20, # AGENCIA
            25, # ALMACENERO
            25, # PROVEEDOR
            20, # FECHA DE ENTRADA
            18, # N° DE RECIBO
            25, # PRODUCTO
            15, # CANTIDAD
        ]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        for entry_item in entry_items:
            ws.append([
                entry_item.entry.agency.name,
                entry_item.entry.warehouse_keeper.first_name + " " + entry_item.entry.warehouse_keeper.last_name,
                entry_item.entry.supplier.name,
                entry_item.entry.entry_date,
                entry_item.entry.invoice_number,
                entry_item.product_stock.product.name,
                entry_item.quantity,
            ])
            
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'inline; filename="reporte_de_entradas_{start_date_str}_a_{end_date_str}.xlsx"'

        wb.save(response)
        return response
    
    
class OutputReportExcelView(APIView):
    """Generate Excel Report."""
    def get(self, request):
        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        try:
            output_items = OutputItem.objects.filter(output__output_date__range=(start_date, end_date)).select_related("output").order_by("output__id")
        except OutputItem.DoesNotExist:
            raise Http404("Salidas no encontradas.")
        wb = Workbook()
        ws = wb.active
        start_date_formatted = start_date.strftime("%d-%m-%y")
        end_date_formatted = end_date.strftime("%d-%m-%y")
        ws.title = f"Salidas {start_date_formatted} - {end_date_formatted}"[:31]
        ws.append(["REPORTE DE SALIDAS DEL " + start_date_formatted + " AL " + end_date_formatted])
        headers = (["AGENCIA",
                   "ALMACENERO",
                   "CLIENTE",
                   "FECHA DE SALIDA",
                   "N° DE RECIBO",
                   "PRODUCTO",
                   "CANTIDAD",
                   ])
        # Unir todas las celdas de la fila 1 (título) y centrarlo
        num_columns = len(headers)
        ws.merge_cells(f'A1:{chr(64 + num_columns)}1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, size=14)
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        column_widths = [
            20, # AGENCIA
            25, # ALMACENERO
            25, # CLIENTE
            20, # FECHA DE ENTRADA
            18, # N° DE RECIBO
            25, # PRODUCTO
            15, # CANTIDAD
        ]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        for output_item in output_items:
            ws.append([
                output_item.output.agency.name,
                output_item.output.warehouse_keeper.first_name + " " + output_item.output.warehouse_keeper.last_name,
                output_item.output.client.name,
                output_item.output.output_date,
                output_item.output.invoice_number,
                output_item.product_stock.product.name,
                output_item.quantity,
            ])
            
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'inline; filename="reporte_de_salidas_{start_date_str}_a_{end_date_str}.xlsx"'

        wb.save(response)
        return response
